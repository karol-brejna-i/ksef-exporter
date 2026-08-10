#!/usr/bin/env python3
"""Export KSeF invoices from the SQLite database to an Excel file.

Usage:
  source .venv/bin/activate
  python scripts/export-invoices.py [--from YYYY-MM-DD] [--to YYYY-MM-DD] [output.xlsx]

Defaults to exporting all invoices and writing `data/invoices-export.xlsx`.
"""

import argparse
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "ksef-exporter.sqlite"
DEFAULT_OUT = Path(__file__).resolve().parent.parent / "data" / "invoices-export.xlsx"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11)
CURRENCY_FORMAT = '#,##0.00'

COLUMNS = [
    ("source", "Źródło", 10),
    ("ksef_number", "Numer KSeF", 38),
    ("invoice_number", "Numer faktury", 24),
    ("issue_date", "Data wystawienia", 14),
    ("seller_nip", "NIP sprzedawcy", 14),
    ("seller_name", "Sprzedawca", 40),
    ("buyer_nip", "NIP nabywcy", 14),
    ("buyer_name", "Nabywca", 40),
    ("gross_total", "Brutto", 12),
    ("currency", "Waluta", 8),
    ("category_name", "Kategoria", 16),
    ("categorization_confidence", "Pewność kategorii", 18),
    ("created_at", "Utworzono", 20),
]


def fetch_invoices(db_path: Path, date_from: str | None = None, date_to: str | None = None) -> list[dict]:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row

    query = """
        SELECT
          i.source,
          i.ksef_number,
          i.invoice_number,
          i.issue_date,
          i.seller_nip,
          i.seller_name,
          i.buyer_nip,
          i.buyer_name,
          i.gross_total,
          i.currency,
          c.name AS category_name,
          i.categorization_confidence,
          i.created_at
        FROM invoices i
        LEFT JOIN categories c ON c.id = i.category_id
    """
    params: list[str] = []
    conditions: list[str] = []

    if date_from:
        conditions.append("i.issue_date >= ?")
        params.append(date_from)
    if date_to:
        conditions.append("i.issue_date <= ?")
        params.append(date_to)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += "\n        ORDER BY i.issue_date DESC, i.invoice_number"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def build_xlsx(invoices: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Faktury"

    # Header row
    for col_idx, (_, label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_idx, inv in enumerate(invoices, start=2):
        for col_idx, (key, _, width) in enumerate(COLUMNS, start=1):
            value = inv.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            if key == "gross_total" and value is not None:
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row & enable auto-filter
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export invoices to Excel")
    parser.add_argument(
        "--from",
        dest="date_from",
        metavar="YYYY-MM-DD",
        help="earliest issue date (inclusive)",
    )
    parser.add_argument(
        "--to",
        dest="date_to",
        metavar="YYYY-MM-DD",
        help="latest issue date (inclusive)",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=str(DEFAULT_OUT),
        help="output .xlsx path (default: data/invoices-export.xlsx)",
    )
    args = parser.parse_args()

    if not DB_PATH.exists():
        print(f"Database not found: {DB_PATH}", file=sys.stderr)
        sys.exit(1)

    invoices = fetch_invoices(DB_PATH, date_from=args.date_from, date_to=args.date_to)
    if not invoices:
        print("No invoices found for the given criteria.")
        sys.exit(0)

    build_xlsx(invoices, Path(args.output))
    print(f"Exported {len(invoices)} invoices → {args.output}")


if __name__ == "__main__":
    main()
